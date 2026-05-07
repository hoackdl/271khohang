from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Exists, OuterRef
from invoice_reader_app.model_invoice import Invoice
from invoice_reader_app.models_purchaseorder import PurchaseOrderItem
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Exists, OuterRef, IntegerField, Value, Case, When
from django.db.models.functions import Cast
from invoice_reader_app.model_invoice import Invoice
# --- Ép số hóa đơn sang số nguyên an toàn --- Danh sách hóa đơn xuất
from django.db.models import Sum, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from .models_purchaseorder import BankPayment, PurchaseOrder
from decimal import Decimal
from django.db.models import Case, When, Value, IntegerField, Q, Exists, OuterRef
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime
from django.db.models import Min, Max
from .model_invoice import Invoice
from invoice_reader_app.model_invoice import Supplier, InvoiceItem, ProductName
from django.shortcuts import render
import openpyxl
from openpyxl.utils import get_column_letter
from django.db.models import Prefetch
from django.db import models






def export_export_orders_excel(request):
    # Lọc dữ liệu theo query params
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    search = request.GET.get("search", "").strip()

    qs = PurchaseOrder.objects.filter(phan_loai_phieu="PX").select_related('invoice')

    if start_date:
        qs = qs.filter(created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(created_at__date__lte=end_date)
    if search:
        qs = qs.filter(
            po_number__icontains=search
        )

    # Tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách PX"

    # Header
    headers = ["PO Number", "Số hóa đơn", "Ngày tạo", "Khách hàng", "MST", "Tổng tiền hóa đơn"]
    ws.append(headers)

    # Data
    for po in qs:
        ws.append([
            po.po_number,
            po.invoice.so_hoa_don if po.invoice else "",
            po.created_at.strftime("%d/%m/%Y"),
            po.invoice.ten_nguoi_mua if po.invoice else "",
            po.invoice.ma_so_thue_mua if po.invoice else "",
            float(po.invoice.tong_tien) if po.invoice else 0
        ])

    # Set column width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"PX_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

def parse_date(date_str):
    """Parse ngày từ 'yyyy-mm-dd' hoặc 'dd/mm/yyyy' sang datetime.date"""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None




from django.http import HttpResponse

def export_invoices_excel(request):
    invoices = Invoice.objects.all()  # Có thể lọc theo GET params giống danh sách

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hóa đơn"

    # Header
    headers = [
        "Số HĐ", "Ngày HĐ", "Tên đơn vị bán", "MST", "Địa chỉ", "HT Thanh toán", "Tổng tiền"
    ]
    ws.append(headers)

    for inv in invoices:
        ws.append([
            inv.so_hoa_don,
            inv.ngay_hd.strftime("%d/%m/%Y") if inv.ngay_hd else "",
            inv.ten_dv_ban,
            inv.ma_so_thue,
            inv.dia_chi,
            inv.hinh_thuc_tt,
            inv.tong_tien
        ])

    # Trả file Excel về client
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=invoices.xlsx'
    wb.save(response)
    return response



from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Exists, OuterRef, IntegerField, Value, Case, When, Sum, DecimalField, Subquery
from django.db.models.functions import Cast, Coalesce



def parse_date(date_str):
    """Parse ngày từ 'yyyy-mm-dd' hoặc 'dd/mm/yyyy' sang datetime.date"""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None



# --- Export Excel ---
def export_invoices_excel(request):
    invoices = Invoice.objects.filter(
        ma_so_thue='0314858906',
        loai_hd='XUAT'
    ).annotate(
        so_hd_int=Case(
            When(so_hoa_don__regex=r'^\d+$', then=Cast('so_hoa_don', IntegerField())),
            default=Value(0),
            output_field=IntegerField()
        ),
        total_paid_amount=Coalesce(
            Subquery(
                BankPayment.objects.filter(
                    purchase_orders__invoice=OuterRef('pk')
                ).values('purchase_orders__invoice').annotate(
                    total=Coalesce(
                        Sum('credit'),
                        Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
                    )
                ).values('total'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            ),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    ).order_by('-so_hd_int', '-ngay_hd')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hóa đơn"

    headers = [
        "Số HĐ", "Ngày HĐ", "Tên đơn vị bán", "MST", "Địa chỉ", "HT Thanh toán",
        "Tổng tiền", "Đã thu", "Tỷ lệ thu (%)"
    ]
    ws.append(headers)

    for inv in invoices:
        total_paid = float(inv.total_paid_amount or 0)
        tong_tien = float(inv.tong_tien or 0)
        percent = round(total_paid / tong_tien * 100, 2) if tong_tien else 0
        ws.append([
            inv.so_hoa_don,
            inv.ngay_hd.strftime("%d/%m/%Y") if inv.ngay_hd else "",
            inv.ten_dv_ban,
            inv.ma_so_thue,
            inv.dia_chi,
            inv.hinh_thuc_tt,
            tong_tien,
            total_paid,
            percent
        ])

    # Set column width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=invoices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def invoice_export_list(request):
    

    # Ưu tiên year trên URL
    year_param = request.GET.get("year")

    if year_param and year_param.isdigit():
        current_year = int(year_param)
        request.session["fiscal_year"] = current_year
    else:
        current_year = request.session.get("fiscal_year", datetime.now().year)


    year_range = (
        Invoice.objects
        .aggregate(
            min_year=Min("fiscal_year"),
            max_year=Max("fiscal_year")
        )
    )

    min_year = year_range["min_year"] or current_year
    max_year = year_range["max_year"] or current_year

    years = set(range(min_year, max_year + 1))
    years.add(current_year)
    year_list = sorted(years)


    # ================= BASE QUERY =================
    invoices = (
        Invoice.objects
        .filter(ma_so_thue='0314858906', loai_hd='XUAT',fiscal_year=current_year)
        .annotate(
            so_hd_int=Case(
                When(
                    so_hoa_don__regex=r'^\d+$',
                    then=Cast('so_hoa_don', IntegerField())
                ),
                default=Value(0),
                output_field=IntegerField()
            )
        )
        .order_by('-ngay_hd', '-so_hd_int')
        .prefetch_related('items')   # ✅ PREFETCH Ở ĐÂY
    )

    # ================= SEARCH =================
    search = request.GET.get('search', '').strip()
    if search:
        invoices = invoices.filter(
            Q(ten_nguoi_mua__icontains=search) |
            Q(ma_so_thue_mua__icontains=search) |
            Q(so_hoa_don__icontains=search) |
            Q(ky_hieu__icontains=search)
        )

    # ================= DATE FILTER =================
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    if start_date and end_date:
        invoices = invoices.filter(ngay_hd__range=(start_date, end_date))
    elif start_date:
        invoices = invoices.filter(ngay_hd__gte=start_date)
    elif end_date:
        invoices = invoices.filter(ngay_hd__lte=end_date)

    # ================= TOTAL PAID =================
    paid_subquery = (
        BankPayment.objects
        .filter(purchase_orders__invoice_id=OuterRef('pk'))
        .values('purchase_orders__invoice_id')
        .annotate(
            total=Coalesce(
                Sum('credit'),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            )
        )
        .values('total')
    )

    invoices = invoices.annotate(
        total_paid_amount=Coalesce(
            Subquery(paid_subquery),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )

    # ================= HAS EXPORTED =================
    invoices = invoices.annotate(
        has_exported=Exists(
            PurchaseOrder.objects.filter(
                invoice_id=OuterRef('pk'),
                phan_loai_phieu='PX'
            )
        )
    )

    # ================= TOTAL ALL =================
    total_tien_all = invoices.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']

    # ================= PAGINATION (LUÔN DÙNG) =================
    per_page_param = request.GET.get('per_page', '10')
    page_number = request.GET.get('page', 1)

    if per_page_param == 'all':
        per_page = invoices.count() or 1   # ✅ trick quan trọng
    else:
        per_page = int(per_page_param) if per_page_param.isdigit() else 10

    # PAGINATION
    paginator = Paginator(invoices, per_page)
    page_number = request.GET.get('page', 1)
    paginated_invoices = paginator.get_page(page_number)

    # ✅ TÍNH TỔNG KHI CÒN LÀ QUERYSET
    total_tien_page = paginated_invoices.object_list.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']

    # 🔥 SAU ĐÓ MỚI ÉP LIST (FIX PREFETCH BUG)
    paginated_invoices.object_list = list(paginated_invoices.object_list)

    for invoice in paginated_invoices.object_list:
        for item in invoice.items.all():
            # Nếu item.sku đã có, dùng luôn; nếu không có, có thể map từ ProductName
            if not item.sku:
                product = ProductName.objects.filter(ten_hang__iexact=item.ten_hang).first()
                item.sku = product.sku if product else ""

    # ================= QUERY PARAMS =================
    query_params = request.GET.copy()
    query_params.pop('page', None)
    
    return render(request, 'invoice_export_list.html', {
        'invoices': paginated_invoices,   # ✅ LUÔN LÀ Page
        'total_tien': total_tien_all,
        'total_tien_page': total_tien_page,
        'search': search,
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'per_page': per_page_param,
        'page_number': paginated_invoices.number,
        'query_params': query_params.urlencode(),
        'current_year': current_year,
        'year_list': year_list,   # 👈 thêm
    })



from django.db.models import (
    Q, Sum, Case, When, Value, IntegerField,
    OuterRef, Subquery, Exists, DecimalField
)
from django.db.models.functions import Cast, Coalesce
from django.core.paginator import Paginator
from django.shortcuts import render
from datetime import datetime
from django.utils.dateparse import parse_date

from invoice_reader_app.model_invoice import Invoice, InvoiceItem
from .models_purchaseorder import BankPayment, PurchaseOrder


def invoice_export_list(request):

    # ================= YEAR =================
    year_param = request.GET.get("year")

    if year_param and year_param.isdigit():
        current_year = int(year_param)
        request.session["fiscal_year"] = current_year
    else:
        current_year = request.session.get("fiscal_year", datetime.now().year)

    year_range = Invoice.objects.aggregate(
        min_year=Min("fiscal_year"),
        max_year=Max("fiscal_year")
    )

    min_year = year_range["min_year"] or current_year
    max_year = year_range["max_year"] or current_year

    years = set(range(min_year, max_year + 1))
    years.add(current_year)
    year_list = sorted(years)

    # ================= BASE QUERY =================
    invoices = (
        Invoice.objects
        .filter(
            ma_so_thue='0314858906',
            loai_hd='XUAT',
            fiscal_year=current_year
        )
        .annotate(
            so_hd_int=Case(
                When(
                    so_hoa_don__regex=r'^\d+$',
                    then=Cast('so_hoa_don', IntegerField())
                ),
                default=Value(0),
                output_field=IntegerField()
            )
        )
        .order_by('-ngay_hd', '-so_hd_int')
        .prefetch_related(
            'items'   # ✅ chỉ load, KHÔNG xử lý ở đây
        )
    )

    # ================= SEARCH =================
    search = request.GET.get('search', '').strip()
    if search:
        invoices = invoices.filter(
            Q(ten_nguoi_mua__icontains=search) |
            Q(ma_so_thue_mua__icontains=search) |
            Q(so_hoa_don__icontains=search) |
            Q(ky_hieu__icontains=search)
        )

    # ================= DATE FILTER =================
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    if start_date and end_date:
        invoices = invoices.filter(ngay_hd__range=(start_date, end_date))
    elif start_date:
        invoices = invoices.filter(ngay_hd__gte=start_date)
    elif end_date:
        invoices = invoices.filter(ngay_hd__lte=end_date)

    # ================= TOTAL PAID =================
    paid_subquery = (
        BankPayment.objects
        .filter(purchase_orders__invoice_id=OuterRef('pk'))
        .values('purchase_orders__invoice_id')
        .annotate(
            total=Coalesce(
                Sum('credit'),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            )
        )
        .values('total')
    )

    invoices = invoices.annotate(
        total_paid_amount=Coalesce(
            Subquery(paid_subquery),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )

    # ================= HAS EXPORTED =================
    invoices = invoices.annotate(
        has_exported=Exists(
            PurchaseOrder.objects.filter(
                invoice_id=OuterRef('pk'),
                phan_loai_phieu='PX'
            )
        )
    )

    # ================= TOTAL ALL =================
    total_tien_all = invoices.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']

    # ================= PAGINATION =================
    per_page_param = request.GET.get('per_page', '10')
    page_number = request.GET.get('page', 1)

    if per_page_param == 'all':
        per_page = invoices.count() or 1
    else:
        per_page = int(per_page_param) if per_page_param.isdigit() else 10

    paginator = Paginator(invoices, per_page)
    paginated_invoices = paginator.get_page(page_number)

    # ✅ tính tổng page khi còn queryset
    total_tien_page = paginated_invoices.object_list.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']

    # 👉 convert sang list để render
    paginated_invoices.object_list = list(paginated_invoices.object_list)

    # ================= FALLBACK SKU (KHÔNG QUERY DB) =================
    for invoice in paginated_invoices.object_list:
        for item in invoice.items.all():
            item.display_sku = item.sku if item.sku else "⚠️ CHƯA MAP"

    # ================= QUERY PARAMS =================
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'invoice_export_list.html', {
        'invoices': paginated_invoices,
        'total_tien': total_tien_all,
        'total_tien_page': total_tien_page,
        'search': search,
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'per_page': per_page_param,
        'page_number': paginated_invoices.number,
        'query_params': query_params.urlencode(),
        'current_year': current_year,
        'year_list': year_list,
    })



from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q, F, Case, When, Value, IntegerField, Subquery, OuterRef, Exists
from django.db.models.functions import Cast, Coalesce
from django.utils.dateparse import parse_date
from decimal import Decimal


def invoice_export_list(request):
    # ================= YEAR =================
    year_param = request.GET.get("year")
    if year_param and year_param.isdigit():
        current_year = int(year_param)
        request.session["fiscal_year"] = current_year
    else:
        from datetime import datetime
        current_year = request.session.get("fiscal_year", datetime.now().year)

    year_range = Invoice.objects.aggregate(
        min_year=models.Min("fiscal_year"),
        max_year=models.Max("fiscal_year")
    )
    min_year = year_range["min_year"] or current_year
    max_year = year_range["max_year"] or current_year
    year_list = sorted(set(range(min_year, max_year + 1)) | {current_year})

    # ================= BASE QUERY =================
    invoices = (
        Invoice.objects
        .filter(ma_so_thue='0314858906', loai_hd='XUAT', fiscal_year=current_year)
        .annotate(
            so_hd_int=Case(
                When(so_hoa_don__regex=r'^\d+$', then=Cast('so_hoa_don', IntegerField())),
                default=Value(0),
                output_field=IntegerField()
            )
        )
        .order_by('-ngay_hd', '-so_hd_int')
        .prefetch_related('items')  # ✅ Prefetch InvoiceItem
    )

    # ================= SEARCH =================
    search = request.GET.get('search', '').strip()
    if search:
        invoices = invoices.filter(
            Q(ten_nguoi_mua__icontains=search) |
            Q(ma_so_thue_mua__icontains=search) |
            Q(so_hoa_don__icontains=search) |
            Q(ky_hieu__icontains=search)
        )

    # ================= DATE FILTER =================
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None
    if start_date and end_date:
        invoices = invoices.filter(ngay_hd__range=(start_date, end_date))
    elif start_date:
        invoices = invoices.filter(ngay_hd__gte=start_date)
    elif end_date:
        invoices = invoices.filter(ngay_hd__lte=end_date)

    # ================= TOTAL PAID =================
    paid_subquery = (
        BankPayment.objects
        .filter(purchase_orders__invoice_id=OuterRef('pk'))
        .values('purchase_orders__invoice_id')
        .annotate(total=Coalesce(Sum('credit'), Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))))
        .values('total')
    )
    invoices = invoices.annotate(
        total_paid_amount=Coalesce(Subquery(paid_subquery, output_field=DecimalField(max_digits=18, decimal_places=2)),
                                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2)))
    )

    # ================= HAS EXPORTED =================
    invoices = invoices.annotate(
        has_exported=Exists(
            PurchaseOrder.objects.filter(invoice_id=OuterRef('pk'), phan_loai_phieu='PX')
        )
    )

    # ================= TOTAL ALL =================
    # Tổng tiền tất cả hóa đơn
    total_tien_all = invoices.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']



    # ================= PAGINATION =================
    per_page_param = request.GET.get('per_page', '10')
    per_page = invoices.count() if per_page_param == 'all' else (int(per_page_param) if per_page_param.isdigit() else 10)
    paginator = Paginator(invoices, per_page)
    page_number = request.GET.get('page', 1)
    paginated_invoices = paginator.get_page(page_number)

    # ================= TOTAL PAGE =================
    # Tổng tiền trên trang hiện tại (pagination)
    total_tien_page = paginated_invoices.object_list.aggregate(
        total=Coalesce(
            Sum('tong_tien'),
            Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
        )
    )['total']

    # ================= USE SKU FROM DB =================
    # Không map lại, chỉ hiển thị từ item.sku đã lưu
    for invoice in paginated_invoices.object_list:
        for item in invoice.items.all():
            item.sku_to_show = item.sku or ""  # tạo thuộc tính hiển thị
    # ================= ADD SKU TO ITEMS =================
    for invoice in paginated_invoices.object_list:
        items_list = list(invoice.items.all())  # convert sang list để có thể gán thuộc tính
        for item in items_list:
            item.sku_to_show = item.sku or ""
        invoice.items_list = items_list  # thêm thuộc tính mới cho invoice
        
    # ================= QUERY PARAMS =================
    query_params = request.GET.copy()
    query_params.pop('page', None)

    return render(request, 'invoice_export_list.html', {
        'invoices': paginated_invoices,
        'total_tien': total_tien_all,
        'total_tien_page': total_tien_page,
        'search': search,
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'per_page': per_page_param,
        'page_number': paginated_invoices.number,
        'query_params': query_params.urlencode(),
        'current_year': current_year,
        'year_list': year_list,
    })