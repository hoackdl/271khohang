import openpyxl
from io import BytesIO
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import HttpResponse
from invoice_reader_app.model_invoice import ProductName
from django.db import models
from .forms import ProductForm


# --- Danh mục hàng hóa (DMHH) ---
def product_dmhh(request):
    search_product = request.GET.get('search_product', '').strip()
    nhom_hang = request.GET.get('nhom_hang', '').strip()

    products_qs = ProductName.objects.all()

    if search_product:
        products_qs = products_qs.filter(
            models.Q(ten_hang__icontains=search_product) |
            models.Q(ten_goi_chung__icontains=search_product) |
            models.Q(sku__icontains=search_product)
        )

    if nhom_hang:
        products_qs = products_qs.filter(nhom_hang=nhom_hang)

    products_qs = products_qs.order_by('id')
 
    # Phân trang: 10 khách hàng/trang
    paginator = Paginator(products_qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Xử lý query params khác ngoài page (để giữ khi chuyển trang)
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')

    context = {
        'products': page_obj,         # trang hiện tại
        'data': page_obj,              # nếu bạn dùng component pagination chung
        'search_product': search_product,
        'query_params': query_params.urlencode(),
        'label': 'hàng hoá',
    }
    return render(request, 'product_dmhh.html', context)


# --- Import Excel ---
import openpyxl
from django.contrib import messages
from django.shortcuts import redirect
import openpyxl
from django.shortcuts import redirect
from django.contrib import messages
from invoice_reader_app.model_invoice import ProductName

def products_import_excel(request):
    """
    Import danh mục hàng hóa từ Excel (.xlsx, .xls)
    - File chỉ có 3 cột: SKU | Tên hàng | Tên gọi chung
    - Nhóm hàng sẽ được gán mặc định nếu file không có cột nhóm
    """
    DEFAULT_NHOM_HANG = 'HH'  # Giá trị mặc định, thay theo NHOM_HANG_CHOICES

    if request.method == 'POST' and request.FILES.get('excel_file'):
        print("Request method:", request.method)
        print("FILES:", request.FILES)

        file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active  # Lấy sheet đầu tiên

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            total_rows = 0
            created_count = 0
            skipped_count = 0

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                print(row)

                sku, ten_hang, ten_goi_chung = row[:3]

                # Bỏ qua dòng nếu thiếu tên hàng
                if not ten_hang or not str(ten_hang).strip():
                    continue

                total_rows += 1

                sku_val = str(sku).strip() if sku else ''
                ten_hang_val = str(ten_hang).strip()
                ten_goi_chung_val = str(ten_goi_chung).strip() if ten_goi_chung else ''

                # Kiểm tra trùng lặp
                if ProductName.objects.filter(
                    sku__iexact=sku_val,
                    ten_hang__iexact=ten_hang_val,
                    ten_goi_chung__iexact=ten_goi_chung_val
                ).exists():
                    skipped_count += 1
                    continue

                # Tạo sản phẩm mới với nhóm hàng mặc định
                ProductName.objects.create(
                    sku=sku_val,
                    ten_hang=ten_hang_val,
                    ten_goi_chung=ten_goi_chung_val,
                    nhom_hang=DEFAULT_NHOM_HANG
                )
                created_count += 1

            messages.success(
                request,
                f"Đã xử lý {total_rows} dòng, tạo mới {created_count}, bỏ qua {skipped_count} trùng."
            )

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f"Lỗi khi import Excel: {str(e)}")

    else:
        messages.error(request, "Vui lòng chọn file Excel trước khi nhập.")

    return redirect('product_dmhh')




# --- Export Excel (tên hàng duy nhất) ---
def products_export_excel(request):
    search_product = request.GET.get('search_product', '').strip()
    products_qs = ProductName.objects.all()
    if search_product:
        products_qs = products_qs.filter(
            models.Q(ten_hang__icontains=search_product) |
            models.Q(ten_goi_chung__icontains=search_product) |
            models.Q(sku__icontains=search_product)
        )

    # Lọc tên hàng duy nhất
    unique_names = {}
    for product in products_qs:
        if product.ten_hang not in unique_names:
            unique_names[product.ten_hang] = product

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh mục hàng hóa"

    # Header
    ws.append(["TT", "SKU", "Tên hàng hóa", "Tên gọi chung"])

    # Dữ liệu
    for idx, product in enumerate(unique_names.values(), start=1):
        ws.append([idx, product.sku, product.ten_hang, product.ten_goi_chung or ''])

    # Lưu file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_dmhh.xlsx'
    return response


def delete_all_products(request):
    if request.method == 'POST':
        count = ProductName.objects.count()
        ProductName.objects.all().delete()
        messages.success(request, f"Đã xóa toàn bộ {count} sản phẩm trong danh mục.")
    else:
        messages.error(request, "Phải gửi POST mới được phép xóa.")
    return redirect('product_dmhh')

def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            sku = form.cleaned_data.get('sku', '').strip()
            ten_hang = form.cleaned_data.get('ten_hang', '').strip()
            ten_goi_chung = form.cleaned_data.get('ten_goi_chung', '').strip()

            # Kiểm tra trùng
            if ProductName.objects.filter(
                sku__iexact=sku,
                ten_hang__iexact=ten_hang,
                ten_goi_chung__iexact=ten_goi_chung
            ).exists():
                messages.warning(request, f"Sản phẩm {ten_hang} đã tồn tại, không thêm được.")
            else:
                ProductName.objects.create(
                    sku=sku,
                    ten_hang=ten_hang,
                    ten_goi_chung=ten_goi_chung
                )
                messages.success(request, f"Đã thêm sản phẩm {ten_hang} thành công.")
            return redirect('product_dmhh')
    else:
        form = ProductForm()

    return render(request, 'product_add_hh.html', {'form': form})


from django.shortcuts import render, get_object_or_404, redirect


def edit_product_dmhh(request, product_id):
    product = get_object_or_404(ProductName, id=product_id)

    if request.method == 'POST':
        product.ten_hang = request.POST.get('ten_hang', product.ten_hang)
        product.sku = request.POST.get('sku', product.sku)
        product.ten_goi_chung = request.POST.get('ten_goi_chung', product.ten_goi_chung)
        # cập nhật các trường khác nếu cần

        product.save()
        messages.success(request, 'Cập nhật sản phẩm thành công.')
        return redirect('inventory_summary')  # chuyển về trang tổng hợp


    return render(request, 'edit_product_dmhh.html', {'product': product})
