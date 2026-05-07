from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from decimal import Decimal
from openpyxl.styles import numbers


from django.db.models import Q
from invoice_reader_app.model_invoice import Invoice

def get_filtered_invoices(request):
    qs = Invoice.objects.prefetch_related("items").all()

    # Năm
    fiscal_year = request.session.get("fiscal_year")
    if fiscal_year:
        qs = qs.filter(fiscal_year=fiscal_year)

    # loại bỏ dữ liệu rác
    qs = qs.exclude(ma_so_thue='0314858906')
    qs = qs.exclude(loai_hd="XUAT")

    # search
    search = request.GET.get("search")
    if search:
        qs = qs.filter(
            Q(ma_so_thue__icontains=search) |
            Q(ten_dv_ban__icontains=search) |
            Q(so_hoa_don__icontains=search)
        )

    # date
    start = request.GET.get("start_date")
    end = request.GET.get("end_date")

    if start:
        qs = qs.filter(ngay_hd__gte=start)
    if end:
        qs = qs.filter(ngay_hd__lte=end)

    return qs.order_by("-ngay_hd")

def export_invoice_input_excel(request):
    ids = request.GET.getlist("ids")

    invoices = Invoice.objects.prefetch_related("items")

    # 👉 nếu có chọn thì lọc theo checkbox
    if ids:
        invoices = invoices.filter(id__in=ids)
    else:
        # fallback: nếu không chọn gì thì dùng filter bình thường
        invoices = get_filtered_invoices(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "HoaDonDauVao"

    ws.append([
        "Tên file",
        "Số HĐ",
        "Ngày HĐ",
        "Nhà cung cấp",
        "MST",
        "Tên hàng",
        "ĐVT",
        "Số lượng",
        "Đơn giá",
        "Thành tiền",
        "Thuế suất",
        "Tiền thuế",
        "Chiết khấu",
        "Thanh toán",
    ])

    for inv in invoices:
        items = list(inv.items.all())

        tong_chiet_khau = sum(i.chiet_khau or 0 for i in items)
        tong_thanh_toan = sum(i.thanh_toan or 0 for i in items)

        first_row = True

        for item in items:
            ws.append([
                inv.file_name if hasattr(inv, "file_name") else "",
                inv.so_hoa_don,
                inv.ngay_hd,
                inv.ten_dv_ban,
                inv.ma_so_thue,

                item.ten_hang,
                item.dvt,
                float(item.so_luong or 0),
                float(item.don_gia or 0),
                float(item.thanh_tien or 0),
                float(item.thue_suat or 0),
                float(item.tien_thue or 0),

                float(tong_chiet_khau) if first_row else "",
                float(tong_thanh_toan) if first_row else "",
            ])

            # format ngày (cột 3)
            ws.cell(row=ws.max_row, column=3).number_format = 'DD/MM/YYYY'

            first_row = False

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="hoa_don_selected.xlsx"'

    wb.save(response)
    return response